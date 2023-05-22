import requests
import numpy as np
import ast
import json
from collections import defaultdict
import sys
import networkx as nx
import os

def load_rep_genes():

    rep_genes = np.loadtxt("DATA/representative_pdb_chain.txt", dtype = str, usecols = 0).tolist()

    rep_genes = ",".join(rep_genes)

    return rep_genes

def get_GO_annotations(gene: str):

    """
    obtain GO annotations from UniProt
    They are leaf nodes 

    REST API: https://www.uniprot.org/help/return_fields

    """

    print(gene)

    UniProt_GO_annot = {}

    gene = gene.strip()

    UniProt_GO_annot[gene] = defaultdict(list)

    UniProt_retrive_tool = requests.get(f"https://rest.uniprot.org/uniprotkb/search?query={gene}&fields=go,id")

    json_uniprot_GO = ast.literal_eval(UniProt_retrive_tool.text)

    if UniProt_retrive_tool.status_code == 200 and json_uniprot_GO["results"]:

        GO_fields = filter(lambda field: field["database"], json_uniprot_GO["results"][0]["uniProtKBCrossReferences"])

        GO_terms = map(lambda field: field["id"], GO_fields)

        for go_term in GO_terms:

            go_id = go_term.split(":")[1]

            go_class_server = requests.get(f"https://www.ebi.ac.uk/QuickGO/services/ontology/go/search?query=GO%3A{go_id}&limit=1&page=1")

            if go_class_server.status_code == 200:

                json_go_class = json.loads(go_class_server.text)

                go_class = map(lambda x: x["aspect"], json_go_class["results"])

                UniProt_GO_annot[gene][list(go_class)[0]].append(go_term)
    
    return UniProt_GO_annot

def run_mp_get_GO_annotations(rep_genes: str):

    import multiprocessing as mp

    cores = len(os.sched_getaffinity(0))
    result_obj = set()

    with mp.get_context("spawn").Pool(cores) as p:
        
        chunk = len(rep_genes.split(",")) // (cores ** 2) + 1

        results = p.map_async(get_GO_annotations, iterable=rep_genes.split(","), chunksize=chunk)
        result_obj.add(results)
        
        for result in result_obj:
            gene_entires = result.get()
    
    UniProt_GO_annot = {}

    for gene_entry in gene_entires:

        gene = list(gene_entry.keys())[0]

        UniProt_GO_annot[gene] = defaultdict(list)

        for go_class, go_terms in gene_entry[gene].items():

            UniProt_GO_annot[gene][go_class].extend(go_terms)
    
    np.savez("DATA/Rep_genes_GO_annot", UniProt_GO_annot)

def transform_leaf_GO_annot(GO_class: str, gene:str, root_node: str, level: int, keyword: str):

    """
    transform leaf nodes for GO terms into their 2nd level
    using is-a relationship for molecular function

    molecular_function: GO:0003674 
    biological_process: GO:0008150
    cellular_component: GO:0005575

    """

    Rep_genes_annot = np.load("DATA/Rep_genes_GO_annot.npz", allow_pickle = True)["arr_0"].tolist()

    GO_class_function = defaultdict(set)

    print(gene)

    for go_term in Rep_genes_annot[gene][GO_class]:

        graph_params = {
            "startIds": go_term, 
            "stopIds": root_node, 
            "relations": keyword}

        graph_url = "https://www.ebi.ac.uk/QuickGO/services/ontology/go/terms/graph"

        graph_response = requests.get(graph_url, params = graph_params)

        if graph_response.status_code == 200:

            json_graph = json.loads(graph_response.text)

            if "results" in json_graph and json_graph["results"]:

                edges = json_graph["results"][0]["edges"]

                vertices = json_graph["results"][0]["vertices"]

                # start_node --> end_node 
                # subject --> object

                graph = nx.DiGraph()

                g_edges = []

                for each_relation in edges:

                    start_node = each_relation["subject"]
                    
                    end_node = each_relation["object"]

                    g_edges.append((start_node, end_node))

                graph.add_edges_from(g_edges)

                toplogical_sorted_groups = list(nx.topological_generations(graph))
                # uses Kahn’s algorithm
                # https://networkx.org/nx-guides/content/algorithms/dag/index.html

                if len(toplogical_sorted_groups) > level: 

                    GO_term_level_in_DAG = toplogical_sorted_groups[-level]

                    GO_labels = [(entry["label"], entry["id"]) for entry in vertices if entry["id"] in GO_term_level_in_DAG]

                    for go_label, entry in GO_labels:

                        GO_class_function[f"{entry}_{go_label}"].add(gene)
                    
    return GO_class_function

def run_mp_transform_leaf(GO_class: str, level: int):

    print(f"RUNNING {GO_class} LEVEL {level}")

    GO_class_root_annot = {
        "molecular_function":["GO:0003674", "is_a"], 
        "biological_process":["GO:0008150", "is_a,part_of"],
        "cellular_component":["GO:0005575", "is_a,part_of,occurs_in"]}

    import multiprocessing as mp

    Rep_genes = list(np.load("DATA/Rep_genes_GO_annot.npz", allow_pickle = True)["arr_0"].tolist())

    params = []

    for rep_gene in Rep_genes:

        params.append((GO_class, rep_gene, GO_class_root_annot[GO_class][0], level, GO_class_root_annot[GO_class][1]))

    cores = len(os.sched_getaffinity(0))
    result_obj = set()

    with mp.get_context("spawn").Pool(cores) as p:
        
        chunk = len(params) // (cores ** 2) + 1

        results = p.starmap_async(transform_leaf_GO_annot, params, chunksize=chunk)
        result_obj.add(results)
        
        for result in result_obj:
            GO_class_transform = result.get()

    GO_class_function = defaultdict(set)

    for entries in GO_class_transform:

        for transform_entry, genes in entries.items():

            if transform_entry:

                GO_class_function[transform_entry].update(genes)

    np.savez(f"DATA/transformed_{level}_lvl_{GO_class}", GO_class_function)

def fishers_exact_test(rep_genes: list, GO_class: str):
  
    """

    Perform a fishers exact test:

    This test deteremines the association between an annotation and entangled genes
    
    Null: There is no association between an annotation and entanglements

    Alt: There is an association between an annotation and entanglements

    2-by-2 contginegency table

                            Have annotation   |  Don"t have annotation
    ----------------------|---------------------------------------------
    | Entangled Genes     |         X         |         A
    ----------------------|---------------------------------------------
    | No Entangled Genes  |         Y         |         B


    """

    from scipy.stats import fisher_exact
    from statsmodels.stats.multitest import fdrcorrection

    transformed_data = np.load(f"DATA/{GO_class}_combined_level_GO_annot.npz", allow_pickle=True)["arr_0"].tolist()

    Ent_genes_PDBs = np.loadtxt("DATA/rep_genes_with_entanglements.txt", dtype = str, usecols = (0, 1))

    Ent_genes = Ent_genes_PDBs[:, 0]

    knotted_proteins = np.load("DATA/percentages_of_knots_in_my_db.npz", allow_pickle=True)["arr_0"].tolist()

    human_knotted_PDBs = np.concatenate([knotted_proteins[organ_kp] for organ_kp in knotted_proteins if organ_kp.startswith("human")])
    
    human_knotted_genes = []

    offset = 0

    for kPDB in human_knotted_PDBs:

        if kPDB in Ent_genes_PDBs[:, 1]:

            idx = np.where(kPDB == Ent_genes_PDBs[:, 1])[0]

            if len(idx) > 1:
                # case in which a PDB has two different genes
                # Ex: 4NDN belong to gene P31153 and gene Q9NZL9; have different chains
                # only in humans
                offset += len(idx) - 1

            human_knotted_genes.extend(Ent_genes_PDBs[idx][:, 0])

    if len(human_knotted_genes) != len(human_knotted_PDBs) + offset:
        print("Did not separate human knotted genes correctly!")
        sys.exit(0)

    disulfide_lassos_from_mapping = np.load("DATA/Disulfide_lassos_from_mapping.npz", allow_pickle = True)["arr_0"].tolist()

    d_ents_genes = [gene.split("_")[1].strip() for gene in disulfide_lassos_from_mapping if gene.startswith("human")]

    Ent_genes = set(Ent_genes) - set(human_knotted_genes) - set(d_ents_genes)

    enriched_results = defaultdict(list)

    all_results = []

    collect_two_tail_pvalues = []

    for transformed_go_term, genes_w_function in transformed_data.items():

        transformed_go_term = transformed_go_term.replace(",", "")
        go_term_no_label = transformed_go_term.split("_")[0]

        if len(genes_w_function) > 1:

            genes_with_no_ent = set(rep_genes) - set(Ent_genes)
            genes_without_function = set(rep_genes) - set(genes_w_function)

            genes_with_ent_and_function = list(set(genes_w_function).intersection(set(Ent_genes))) # X
            genes_with_no_ent_and_function = list(set(genes_w_function).intersection(set(genes_with_no_ent))) # Y

            genes_with_ent_and_no_function = list(set(genes_without_function).intersection(set(Ent_genes))) # A
            genes_with_no_ent_and_no_function = list(set(genes_without_function).intersection(set(genes_with_no_ent))) # B
            
            table = [[len(genes_with_ent_and_function), len(genes_with_ent_and_no_function)], 
                    [len(genes_with_no_ent_and_function), len(genes_with_no_ent_and_no_function)]]
            
            OR, p_value_e = fisher_exact(table, alternative = "greater") # right

            _, p_value_d = fisher_exact(table, alternative = "less") # left

            _, p_value_two_tails = fisher_exact(table)

            if p_value_e <= 0.05 and p_value_d <= 0.05: 
                print("Something's wrong with left and right p_values")
                sys.exit(0)

            all_results.append( [transformed_go_term, OR, p_value_e, p_value_d, 
                p_value_two_tails, len(genes_with_ent_and_function), len(genes_with_ent_and_no_function), 
                len(genes_with_no_ent_and_function), len(genes_with_no_ent_and_no_function), len(genes_w_function), 
                len(genes_without_function)] )

            collect_two_tail_pvalues.append(p_value_two_tails)
    
    rejects, adjusted_two_tail = fdrcorrection(collect_two_tail_pvalues, alpha = 0.05)
    # rejects is true if null hypothesis is rejected or False otherwise

    for result, q_value in zip(all_results, adjusted_two_tail):

        _, _, e_pvalue, d_pvalue, _, _, _, _, _, _, _ = result

        result.insert(5, q_value)

        result.append("True") if q_value <= 0.05 else result.append("False")

    return all_results

def combine_GO_annotations_lvl_wise():

    """
    combine level 2 and 3 GO annotations since there are some overlap

    """

    for go_class in ["molecular_function", "biological_process", "cellular_component"]:

        merged_information = defaultdict(set)

        second_lvl = np.load(f"DATA/transformed_2_lvl_{go_class}.npz", allow_pickle=True)["arr_0"].tolist()

        third_lvl = np.load(f"DATA/transformed_3_lvl_{go_class}.npz", allow_pickle=True)["arr_0"].tolist()

        for each_annot in second_lvl: 

            merged_information[each_annot].update(second_lvl[each_annot])

        for each_annot in third_lvl: 

            merged_information[each_annot].update(third_lvl[each_annot])

        np.savez(f"DATA/{go_class}_combined_level_GO_annot", merged_information)

def create_excel(rep_genes: list):

    from openpyxl import Workbook

    wb = Workbook()
    wb["Sheet"].title = "Human_molecular_function"

    wb.create_sheet("Human_biological_process")
    wb.create_sheet("Human_cellular_component")

    for go_class in ["molecular_function", "biological_process", "cellular_component"]:

        all_results = []

        print(f"PERFORMING FISHER'S EXACT TEST FOR: {go_class}")

        all_lvl_results = fishers_exact_test(rep_genes, go_class)

        sheet = wb[f"Human_{go_class}"]

        headers = ["GO_label", "Odd Ratio", "Enriched_pvalue", 
                    "Depleted_pvalue", "Two_tail_pvalue", "Adj_Two_tail_pvalue", "Ent_gene_w_AO", 
                    "Ent_gene_wo_AO", "Non_ent_genes_w_AO", "Non_ent_genes_wo_AO",
                    "Rep_gene_w_AO", "Rep_gene_wo_AO", "Significant_or_Not"]

        column = 1

        for row_label in headers:

            sheet.cell(1, column).value = row_label

            column += 1

        cell_row = 2

        for rowitems in all_lvl_results:

            cell_column = 1

            for item in rowitems:

                sheet.cell(cell_row, cell_column).value = item
                cell_column += 1

            cell_row += 1
        
    wb.save("DATA/Human_GO_stats.xlsx")

    wb.close()


if __name__ == "__main__":

    rep_genes = load_rep_genes()

    # run_mp_get_GO_annotations(rep_genes)

    # for go_class in ["molecular_function", "biological_process", "cellular_component"]:

    #     run_mp_transform_leaf(go_class, 2)
    #     run_mp_transform_leaf(go_class, 3)

    # combine_GO_annotations_lvl_wise()

    create_excel(rep_genes.split(","))


